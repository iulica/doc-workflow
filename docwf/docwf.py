""" process_bills - process bill pdfs through different tasks """

import sys
import json
import pathlib
import importlib
import copy
from typing import Optional

from PyPDF2 import PdfReader, PdfWriter

# import cairosvg
from openpyxl import load_workbook

from .plugins.base import BasePlugin
from .column_parser import get_parser_map

PROCESS_BILLS_CONFIG = "docwf.json"

class PluginManager():
    """ Plugin Manager
    it holds a cache of task plugins
    """

    def __init__(self, plugin_dir="docwf.plugins", plugins=None):
        """ Load plugins from plugin_dir but external plugins can be provided by the plugin_classes dictionary
        """
        self.plugin_dir = plugin_dir
        self.plugins = {} # name -> module
        if plugins:
            self.plugins.update(plugins)

    def reload(self):
        """ reload the plugins """
        # TODO check if this works or one needs to unload the module with importlib
        self.plugins.clear()

    def get_plugin_class(self, plugin_name):
        """ returns the plugin class for the plugin name """
        plugin_module = self.plugins.get(plugin_name)
        if plugin_module is None:
            plugin_module = self.load_plugin(plugin_name)
            if not plugin_module:
                return None
            self.plugins[plugin_name] = plugin_module
        return plugin_module.PluginClass

    def load_plugin(self, plugin_name):
        """ tries to load a plugin by name """
        try:
            return importlib.import_module("." + plugin_name, package=self.plugin_dir)
        except ImportError:
            raise
            return None

    def get_task(self, name, *args, **kwargs) -> Optional[BasePlugin]:
        """ gets a request object by creating a new object with
         the given parameters for the named plugin"""
        plugin_class = self.get_plugin_class(name)
        if plugin_class:
            return plugin_class(*args, **kwargs)

class InputPDFWrapper():
    def __init__(self, input_filename):
        self.pdf_reader = PdfReader(input_filename)
        self._crt_page = 0
    
    def readPages(self, no_pages):
        pages = [
            self.pdf_reader.pages[page_index]
            for page_index in range(self._crt_page, self._crt_page + no_pages)
        ]
        self._crt_page += no_pages
        return pages
    
    def reset(self):
        self._crt_page = 0
    
    def close(self):
        self.pdf_reader.stream.close()

class TaskHelper():

    """ class used to help with the tasks input and output """

    def __init__(self, task_definition):
        self._makedir_name = task_definition.get('makedir')
        self._input_filename = task_definition.get('input_filename')
        self._output_filename = task_definition.get('output_filename')
        self._delete_input = task_definition.get('delete_input', False)
        self._input_files = {} # map of input file name to the input pdf file
        self._output_files = {} # map of output file names to output streams
        self._made_folders = set()
        self.pages_per_bill = task_definition.get('pages')

    def get_io_streams(self, value_dict):
        output_stream = input_stream = None
        if self._output_filename is not None:
            output_filename = self._output_filename.format(**value_dict)
            output_stream = self._output_files.get(output_filename)
            if output_stream is None:
                output_stream = PdfWriter()
                self._output_files[output_filename] = output_stream

        if self._input_filename is not None:
            input_filename = self._input_filename.format(**value_dict)
            input_stream = self._input_files.get(input_filename)
            if input_stream is None:
                input_stream = self._input_files[input_filename] = InputPDFWrapper(input_filename)

        return input_stream, output_stream

    def _makedir(self, value_dict):
        if self._makedir_name is not None:
            makedir_name = self._makedir_name.format(**value_dict)
            if makedir_name not in self._made_folders:
                self._made_folders.add(makedir_name)
                pathlib.Path(makedir_name).mkdir(parents=True, exist_ok=True)
    
    def do_task(self, value_dict):
        self._makedir(value_dict) # create directory if needed

    def finish(self):
        for output_filename, pdf_writer in self._output_files.items():
            with open(output_filename, "wb") as output_stream:
                pdf_writer.write(output_stream)
        self._output_files.clear()
        for input_filename, input_pdf_wrapper in self._input_files.items():
            input_pdf_wrapper.close()
            if self._delete_input:
                pathlib.Path(input_filename).unlink()
        self._input_files.clear()


class DocWorkflow():
    """ class handling bill processes 
    
        The configuration comes from a dictionary
        Recognized keys:
        "globals":{
            # global configuration
            #  key : value
            #  key : dict
            "workbook": "", # workbook file name
            "sheet": "", # workbook sheet to load
            "constants":{} # constant attributes to add to the task params
        }
        The Global configuration can be 
        smartly (when dict, overwride only the specific keys) overwidden
        through "locals" dictionary in each task definition

        "tasks": [
            {
                "name": "task name", # just for pretty printing
                "active": 0/1, # if the task is active or not
                "params":{ # parameters of each row, gotten from the sheet rows
                    key: value,
                    key: {transformation},
                },
                "filter": {filter definiton}, # a filter to allow the row or skip it
                "task": { # task definition
                    "type": "type of task", # important to know which plugin to use
                    "input_filename": "bills/rechnungen_{key_year}/{key_email}/R{key_year}_{key_house}.pdf", # will be used for a task helper
                    "pages": 2, # number of pages per bill in the input file
                    "output_filename": "bills/rechnungen_{key_year}/{key_email}/R{key_year}_{key_house}_MUSTER.pdf" # buffering the output, will be saved upon finishing the task
                    task specific definition
                }
            }
        ]

    """

    def __init__(self, config, plugins=None, **kwargs):
        self.config = config
        self.workbook_cache = {}
        self.plugin_manager = PluginManager(plugins=plugins)

    def _make_filter_function(self, filter_dict):
        return lambda x:x.get(filter_dict['column']) == filter_dict['value']

    def _build_parser_map(self, column_map_definition, sheet):

        column_index_map = {
            cell.value.strip().lower(): cell.column - 1
            for cell in sheet[1]
            if cell.value
        }
        return get_parser_map(column_map_definition, column_index_map)

    def _get_workbook(self, workbook_name):
        workbook = self.workbook_cache.get(workbook_name)
        if workbook is None:
            workbook = load_workbook(
                filename=workbook_name,
                data_only=True)
            # print(workbook.sheetnames)
            self.workbook_cache[workbook_name] = workbook
        return workbook

    def _create_task(self, task_type, *args, **kwargs):
        return self.plugin_manager.get_task(task_type, *args, **kwargs)

    def process_task(self, task_info, globals_dict):

        task_name = task_info.get('name', 'N/A')
        print("Doing task:", task_name)
        globals_dict_update = task_info.get('locals')
        if globals_dict_update:
            globals_dict = copy.deepcopy(globals_dict)
            for key, value in globals_dict_update.items():
                if isinstance(value, dict):
                    globals_dict.setdefault(key, {}).update(value)
                else:
                    globals_dict[key] = value

        workbook = self._get_workbook(globals_dict['workbook'])
        sheet = workbook[globals_dict['sheet']]

        task_params = globals_dict.get('constants', {}).copy()
        task_local_params = globals_dict.get('task_params', {}).copy()
        # task_local_params.update(task_info.get('task_params', {}))
        # print(task_local_params)
        parser_map = self._build_parser_map(task_local_params, sheet)

        filter_func = None
        if task_info.get('filter'):
            filter_func = self._make_filter_function(task_info['filter'])

        task_dict = task_info['task']
        task_helper = TaskHelper(task_dict)
        task_handler = self._create_task(
            task_dict['type'],
            task_helper, task_dict, globals_dict,
            sheet=sheet)
        if task_handler is None:
            raise RuntimeError("Missing plugin {}".format(task_dict['type']))
        # task_handler = SendEmailTask(globals_dict, task_info['output'])
        for index in range(2, sheet.max_row+1):
            row = sheet[index]
            value_dict = {
                output_name: parser.get_value(row)
                for output_name, parser in parser_map.items()
            }
            if not any(value_dict.values()):
                # print("ignore empty row", index)
                continue
            # print(value_dict)
            task_params.update(value_dict)
            if filter_func and not filter_func(task_params):
                # print("fitered out")
                continue

            task_helper.do_task(task_params)
            task_handler.do_task(task_params, index=index, row=row)

        task_handler.finish()
        task_helper.finish()

    def gen(self):
        globals_dict = self.config.get('globals', {})
        for task_info in self.config['tasks']:
            if not task_info.get('active', 1):
                continue
            self.process_task(task_info, globals_dict)

    @classmethod
    def main(cls, config_file_name, **kwargs):
        config_obj = json.loads(open(config_file_name, 'r', encoding='utf-8').read())
        cls(config_obj, **kwargs).gen()
    
    @staticmethod
    def cli():
        config_file_name = PROCESS_BILLS_CONFIG
        if sys.argv[1:]:
            config_file_name = sys.argv[1]
        DocWorkflow.main(config_file_name)


if __name__ == "__main__":
    DocWorkflow.cli()