""" email plugin - sends emails with bills """

from .base import BaseDataPlugin

class ExcelWorkbook(BaseDataPlugin):

    """ class used to read excel workbooks """

    def load(self):
        try:
            from openpyxl import load_workbook
        except ModuleNotFoundError:
            # install openpyxl to use Excel data type
            raise
        self._workbook = load_workbook(
            filename=self._data_dict['workbook'],
            data_only=True)
        return self

    def get_column_index_map(self, sheet):
        column_index_map = {
            cell.value.strip().lower(): cell.column - 1
            for cell in sheet[1]
            if cell.value
        }
        return column_index_map
        
    def get_value(self, sheet, row, column_index):
        return row[column_index].value

    def get_worksheets(self):
        """
        Returns a list of worksheets
        """

        pass

    def get_worksheet(self, name):
        """returns a worksheet by name"""
        return self._workbook[name]

    def iterrows(self, sheet):
        """iterates over the rows and gives back a index, dictionary column:value pairs"""
        for index in range(2, sheet.max_row+1):
            row = sheet[index]
            yield (index, row)


PluginClass = ExcelWorkbook